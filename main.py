import os
import io
import json
import sys
from openpyxl.reader.excel import load_workbook

LANG = 'lang'
FILTOSI18N = 'filt-os-i18n'
TIMEZONE = 'timezone'

def app_path():
    if hasattr(sys, 'frozen'):
        return os.path.dirname(sys.executable)  # 使用pyinstaller打包后的exe目录
    return os.path.dirname(__file__)  # 没打包前的py目录

def excel2json(path, excel_file):
    wb = load_workbook(excel_file)  # 加载excel表格
    filename = os.path.splitext(os.path.basename(excel_file))[0]
    worksheets_len = len(wb.worksheets)
    for sheet in wb.worksheets:  # wb.worksheets: 获取所有工作表对象
        result = {}
        lang_list = []
        # 生成语种列表
        for column in range(sheet.max_column):
            # 排除 key
            if column > 0 and sheet.cell(1, column + 1).value:
                lang = sheet.cell(1, column + 1).value
                lang_list.append(lang)
                result[lang] = {}

        for row in range(sheet.max_row):  # sheet.max_row: 获取行数
            # 排除表头
            if row > 0:
                for lang in lang_list:
                    key = sheet.cell(row + 1, 1).value
                    if key:
                        val = sheet.cell(row + 1, lang_list.index(lang) + 2).value
                        result[lang][key] = val or key
        save_json_file(path, filename, worksheets_len, sheet.title, lang_list, result)
    wb.close()

def save_json_file(path, filename, worksheets_len, customer, lang_list, result):
    if worksheets_len > 1:
        customer_dir = f"{path}\\{filename}\\{customer}"
    elif filename != TIMEZONE:
        customer_dir = f"{path}\\{filename}"
    else:
        customer_dir = f"{path}"
    if not os.path.exists(customer_dir):
        os.makedirs(customer_dir)
    for lang in lang_list:
        file_path = f"{customer_dir}\\{lang}.json"
        file = io.open(file_path, 'w', encoding='utf-8')
        # 把对象转化为json对象
        # indent: 参数根据数据格式缩进显示，读起来更加清晰
        # ensure_ascii = True：默认输出ASCII码，如果把这个改成False, 就可以输出中文。
        txt = json.dumps(result[lang], indent=2, ensure_ascii=False)
        file.write(txt)
        file.close()
        print(f"已输出{file_path}")

if __name__ == '__main__':
    excels_needed = [f"{LANG}.xlsx", f"{FILTOSI18N}.xlsx", f"{TIMEZONE}.xlsx"]
    excels = []
    files_path = input("请输入翻译文件所在的目录（支持相对路径，不填默认当前目录）：")
    trans_path = input("请输入翻译文件输出路径（不填默认当前目录下的trans文件夹）：")
    if not files_path.strip():
        files_path = '.\\'
    if not trans_path:
        trans_path = '.\\trans'
    try:
        # 遍历指定目录
        for filename in os.listdir(files_path):
            if filename.endswith('.xlsx') and filename in excels_needed:
                excel_path = os.path.join(files_path, filename)
                excels.append(excel_path)
    except FileNotFoundError:
        print("错误: 指定的目录未找到!")
    except Exception as e:
        print(f"错误: 发生了一个未知错误: {e}")
    app_path()
    for excel in excels:
        excel2json(trans_path, excel)
    print("输出完成")
    input('Press Enter to exit...')
